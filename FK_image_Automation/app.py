import os
import shutil
from flask import Flask, request, render_template, send_file
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

app = Flask(__name__)

# Folders
UPLOAD_FOLDER = 'uploads'
PROCESSED_FOLDER = 'processed'
MAIN_IMAGE_FOLDER = 'Mainimage'
SCREENSHOT_FOLDER = 'fullProduct_screenshot'

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)
os.makedirs(MAIN_IMAGE_FOLDER, exist_ok=True)
os.makedirs(SCREENSHOT_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def create_dummy_image(path, color):
    """Creates a valid 100x100 PNG image"""
    img = PILImage.new('RGB', (100, 100), color=color)
    img.save(path, 'PNG')
    print(f"Created dummy image: {path}")

def generate_dummy_images(fsn_list):
    """Generate dummy images for testing"""
    for fsn in fsn_list:
        if not fsn: continue
        fsn = str(fsn).strip()
        
        main_path = os.path.join(MAIN_IMAGE_FOLDER, f"{fsn}_main.png")
        screen_path = os.path.join(SCREENSHOT_FOLDER, f"{fsn}_screen.png")
        
        create_dummy_image(main_path, (255, 0, 0))    # Red
        create_dummy_image(screen_path, (0, 0, 255))   # Blue

def process_excel(file_path):
    print(f"Processing file: {file_path}")
    
    try:
        # Load workbook
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        print(f"Loaded sheet: {ws.title}")
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return None

    # Get FSNs from Column D
    fsn_list = []
    for row in range(2, ws.max_row + 1):
        val = ws[f"D{row}"].value
        if val:
            fsn_list.append(str(val).strip())
    
    print(f"Found {len(fsn_list)} FSNs: {fsn_list}")

    if not fsn_list:
        print("No FSNs found. Saving empty file.")
    else:
        # Generate images
        generate_dummy_images(fsn_list)

        # Insert Images
        for row in range(2, ws.max_row + 1):
            fsn = str(ws[f"D{row}"].value).strip()
            if not fsn: continue

            # --- Column T (Index 20) ---
            main_path = os.path.join(MAIN_IMAGE_FOLDER, f"{fsn}_main.png")
            if os.path.exists(main_path):
                try:
                    img = Image(main_path)
                    img.width = 100
                    img.height = 100
                    ws[f"T{row}"] = img
                    print(f"Inserted main image for {fsn} at T{row}")
                except Exception as e:
                    print(f"Error inserting main image {fsn}: {e}")

            # --- Column U (Index 21) ---
            screen_path = os.path.join(SCREENSHOT_FOLDER, f"{fsn}_screen.png")
            if os.path.exists(screen_path):
                try:
                    img = Image(screen_path)
                    img.width = 150
                    img.height = 150
                    ws[f"U{row}"] = img
                    print(f"Inserted screenshot for {fsn} at U{row}")
                except Exception as e:
                    print(f"Error inserting screenshot {fsn}: {e}")

    # --- CRITICAL FIX: Set Column Widths ---
    # Columns T and U need to be wide enough to show 100px+ images
    # 1 unit of Excel column width is approx 7 pixels. 
    # To show 100px, we need width ~15 (15 * 7 = 105)
    ws.column_dimensions['T'].width = 18
    ws.column_dimensions['U'].width = 24
    
    # Also ensure row height is enough (default is 15, we need ~75 for 100px image)
    # We set row height for all rows that have images
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 80 

    # Save as .xlsx
    base_name = os.path.basename(file_path)
    if '.' in base_name:
        base_name = base_name.rsplit('.', 1)
    
    output_filename = f"processed_{base_name}.xlsx"
    output_path = os.path.join(PROCESSED_FOLDER, output_filename)
    
    try:
        wb.save(output_path)
        print(f"Saved successfully: {output_path}")
        return output_filename
    except Exception as e:
        print(f"Error saving file: {e}")
        return None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return "No file part", 400
    
    file = request.files['file']
    if file.filename == '':
        return "No selected file", 400

    if file:
        filename = file.filename
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        output_filename = process_excel(file_path)
        
        if not output_filename:
            return "Processing failed", 500

        output_path = os.path.join(PROCESSED_FOLDER, output_filename)
        if not os.path.exists(output_path):
            return "File not generated", 500

        return send_file(output_path, as_attachment=True, download_name=output_filename)

if __name__ == '__main__':
    app.run(debug=True, port=5000)